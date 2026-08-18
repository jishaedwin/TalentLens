"""
TalentLens backend — Excel report generation.

Structured .xlsx export of a completed screening's candidate ranking, for
recruiters who want to do further analysis outside the app.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BAND_FILLS = {
    "Strong Fit": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "High Potential": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "Needs Review": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "Low Fit": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}
INTEGRITY_LABELS = {
    "CLEAR": "Verified",
    "WARNING": "Review Required",
    "POTENTIAL MANIPULATION": "Potential Manipulation",
    "UNKNOWN": "Not Evaluated",
}

COLUMNS = [
    ("Rank", 6), ("Candidate Name", 28), ("Resume ID", 14), ("Match Score", 12),
    ("Category", 16), ("Matched Skills", 40), ("Missing Skills", 30),
    ("Experience (yrs)", 14), ("Qualification", 24), ("Integrity Status", 20),
    ("Key Explanation", 60),
]


def generate_excel_report(screening: dict) -> bytes:
    candidates = sorted(screening["candidates"], key=lambda c: c["composite_score"], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Ranking"

    ws.merge_cells("A1:K1")
    ws["A1"] = f"TalentLens — {screening['jd_dict']['job_title']} — Screening Results"
    ws["A1"].font = Font(bold=True, size=13, color="111827")
    ws.row_dimensions[1].height = 24

    header_row = 3
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for i, c in enumerate(candidates, 1):
        row = header_row + i
        values = [
            i,
            c["headline"] or "Untitled",
            c["resume_id"],
            round(c["composite_score"], 1),
            c["band"],
            ", ".join(c["matched_skills"]),
            ", ".join(c["missing_skills"]),
            round(c["years_experience"], 1) if c["years_experience"] else None,
            ", ".join(c.get("education", [])) or "n/a",
            INTEGRITY_LABELS.get(c.get("integrity_status"), "Not Evaluated"),
            c.get("explanation", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            band_fill = BAND_FILLS.get(c["band"])
            if col_idx == 5 and band_fill:
                cell.fill = band_fill

    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
